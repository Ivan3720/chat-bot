import json
import os
import random
import re
import shutil
import tempfile
import zipfile
from datetime import datetime
from urllib.parse import urlparse

import requests
from dotenv import load_dotenv
import vk_api
from vk_api import VkUpload
from vk_api.longpoll import VkEventType, VkLongPoll

# Импорт функций из соседних файлов
from main_menu import create_folders_keyboard, start_keyboard
from tech_cards import get_subfolders, load_all_cards
from tests_data import DESSERT_TESTS, DRINK_TESTS

# Загружаем локальные переменные окружения из .env.
# На сервере эти же значения можно задать через systemd/environment.
load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_CARDS_DIR = os.path.abspath(os.path.join(BASE_DIR, "cards"))
EXAMPLE_FILES_DIR = os.path.join(BASE_DIR, "tech_card_examples")
EXAMPLE_TECH_CARD_FILENAME = "Смузи клубника-банан 0,3.xlsx"

TOKEN = os.getenv("VK_TOKEN")
if not TOKEN:
    raise RuntimeError(
        "Не задан VK_TOKEN. Создайте .env по примеру .env.example "
        "или задайте переменную окружения VK_TOKEN на сервере."
    )

# Секретные значения обязательно меняйте перед сервером.
ADMIN_SECRET = os.getenv("ADMIN_SECRET")
if not ADMIN_SECRET:
    raise RuntimeError(
        "Не задан ADMIN_SECRET. Создайте .env по примеру .env.example "
        "или задайте переменную окружения ADMIN_SECRET на сервере."
    )

ADMIN_ACCESS_WORD = os.getenv("ADMIN_ACCESS_WORD", "admin_007").strip().lower()
ADMIN_IDS_FILE = os.path.join(BASE_DIR, "admins.json")
ALLOWED_CARD_EXTENSIONS = {".xlsx"}
MAX_UPLOAD_SIZE = int(os.getenv("MAX_UPLOAD_SIZE", str(20 * 1024 * 1024)))

# Хранилище ID сообщений для очистки чата
navigation_history = {}
user_pages = {}
user_test_sessions = {}
admin_states = {}


def clear_nav(vk, user_id):
    """Удаляет старые навигационные сообщения для чистоты диалога."""
    if user_id in navigation_history and navigation_history[user_id]:
        try:
            vk.messages.delete(
                message_ids=navigation_history[user_id],
                delete_for_all=1,
                peer_id=user_id,
            )
            navigation_history[user_id] = []
        except Exception:
            pass


def send_msg(vk, user_id, message, keyboard=None, is_final=False, attachment=None):
    """Отправляет сообщение и записывает его ID для последующего удаления."""
    if is_final:
        clear_nav(vk, user_id)

    params = {
        "user_id": user_id,
        "message": message,
        "random_id": random.randint(0, 2**31),
    }

    if keyboard is not None:
        params["keyboard"] = keyboard
    if attachment is not None:
        params["attachment"] = attachment

    msg_id = vk.messages.send(**params)

    if not is_final:
        navigation_history.setdefault(user_id, []).append(msg_id)


def _chunked(items, size):
    """Разбивает список на строки клавиатуры."""
    return [items[i : i + size] for i in range(0, len(items), size)]


def _split_text_blocks(lines, header="", max_len=3500):
    """Разбивает длинный список на несколько сообщений для лимита VK."""
    blocks = []
    current = header.strip()

    for line in lines:
        candidate = f"{current}\n{line}" if current else line
        if len(candidate) > max_len and current:
            blocks.append(current)
            current = line
        else:
            current = candidate

    if current:
        blocks.append(current)

    return blocks


# =========================
# АДМИНИСТРАТИВНАЯ ЧАСТЬ
# =========================


def load_admin_ids():
    """Загружает список администраторов из admins.json и переменной ADMIN_IDS."""
    ids = set()

    env_ids = os.getenv("ADMIN_IDS", "")
    for raw_id in env_ids.split(","):
        raw_id = raw_id.strip()
        if raw_id.isdigit():
            ids.add(int(raw_id))

    if os.path.exists(ADMIN_IDS_FILE):
        try:
            with open(ADMIN_IDS_FILE, "r", encoding="utf-8") as file:
                data = json.load(file)
            for admin_id in data.get("admins", []):
                if str(admin_id).isdigit():
                    ids.add(int(admin_id))
        except Exception as error:
            print(f"Не удалось прочитать {ADMIN_IDS_FILE}: {error}")

    return ids


def save_admin_ids(admin_ids):
    """Сохраняет список администраторов в admins.json."""
    with open(ADMIN_IDS_FILE, "w", encoding="utf-8") as file:
        json.dump({"admins": sorted(admin_ids)}, file, ensure_ascii=False, indent=2)


def is_admin(user_id, admin_ids):
    return int(user_id) in admin_ids


def add_admin(user_id, admin_ids):
    admin_ids.add(int(user_id))
    save_admin_ids(admin_ids)


def create_start_keyboard_for_user(user_id, admin_ids):
    """Главная клавиатура. Админу добавляем кнопку админ-панели."""
    buttons = [
        [{"action": {"type": "text", "label": "📖 Меню"}, "color": "primary"}],
        [{"action": {"type": "text", "label": "📝 Тесты"}, "color": "primary"}],
        [
            {
                "action": {"type": "text", "label": "ℹ️ Общая информация"},
                "color": "primary",
            }
        ],
    ]
    if is_admin(user_id, admin_ids):
        buttons.append(
            [{"action": {"type": "text", "label": "⚙️ Админка"}, "color": "positive"}]
        )
    return json.dumps({"one_time": False, "buttons": buttons}, ensure_ascii=False)


def create_admin_keyboard():
    """Главное меню администратора."""
    buttons = [
        [
            {"action": {"type": "text", "label": "➕ Добавить техкарту"}, "color": "positive"},
            {"action": {"type": "text", "label": "🗑 Удалить техкарту"}, "color": "negative"},
        ],
        [
            {"action": {"type": "text", "label": "📂 Создать папку"}, "color": "positive"},
            {"action": {"type": "text", "label": "🗂 Удалить папку"}, "color": "negative"},
        ],
        [
            {"action": {"type": "text", "label": "📘 Как оформить техкарту"}, "color": "primary"},
            {"action": {"type": "text", "label": "📁 Папки техкарт"}, "color": "primary"},
        ],
        [
            {"action": {"type": "text", "label": "📦 Скачать все техкарты ZIP"}, "color": "secondary"},
            {"action": {"type": "text", "label": "🔄 Обновить базу"}, "color": "primary"},
        ],
        [
            {"action": {"type": "text", "label": "👥 Список админов"}, "color": "secondary"},
        ],
        [
            {"action": {"type": "text", "label": "🏠 В начало"}, "color": "secondary"},
        ],
    ]
    return json.dumps({"one_time": False, "buttons": buttons}, ensure_ascii=False)


def create_admin_folders_keyboard(folders, current_path, mode):
    """Клавиатура выбора папки для загрузки/удаления/создания."""
    buttons = []
    if mode in {"create", "delete_dir"}:
        folders_limit = 9 if current_path == "cards" else 8
    else:
        folders_limit = 8
    for folder in folders[:folders_limit]:
        buttons.append(
            [{"action": {"type": "text", "label": folder[:40]}, "color": "positive"}]
        )

    # Для загрузки файл можно положить даже в текущую папку, если она подходит.
    if current_path != "cards":
        if mode == "upload":
            buttons.append(
                [
                    {
                        "action": {"type": "text", "label": "✅ Выбрать эту папку"},
                        "color": "primary",
                    }
                ]
            )
        if mode == "delete":
            buttons.append(
                [
                    {
                        "action": {"type": "text", "label": "📄 Показать файлы здесь"},
                        "color": "primary",
                    }
                ]
            )
        if mode == "delete_dir":
            buttons.append(
                [
                    {
                        "action": {"type": "text", "label": "🗑 Удалить эту папку"},
                        "color": "negative",
                    }
                ]
            )

    if mode == "create":
        buttons.append(
            [
                {
                    "action": {"type": "text", "label": "➕ Создать здесь"},
                    "color": "primary",
                }
            ]
        )

    buttons.append(
        [
            {"action": {"type": "text", "label": "⬅️ Админ назад"}, "color": "secondary"},
            {"action": {"type": "text", "label": "⚙️ Админка"}, "color": "secondary"},
        ]
    )
    return json.dumps({"one_time": False, "buttons": buttons[:10]}, ensure_ascii=False)


def get_admin_subfolders(path="cards"):
    """Возвращает реальные подпапки для админки.

    Здесь специально не используем обычный get_subfolders(), потому что он
    скрывает некоторые разделы в пользовательском меню. В админке нужно видеть
    все папки, даже пустые, чтобы их можно было удалить или выбрать для загрузки.
    """
    path = ensure_safe_cards_path(path)
    if not os.path.exists(path):
        return []

    return sorted(
        name
        for name in os.listdir(path)
        if os.path.isdir(os.path.join(path, name)) and name != "__MACOSX"
    )


def sanitize_filename(filename):
    """Убирает опасные символы из имени загружаемого файла."""
    filename = os.path.basename(filename)
    filename = re.sub(r"[\\/:*?\"<>|]+", "_", filename)
    filename = filename.strip()
    return filename or "tech_card.xlsx"

def sanitize_folder_name(folder_name):
    """Убирает опасные символы из названия создаваемой папки."""
    folder_name = str(folder_name).strip()
    folder_name = re.sub(r"[\\/:*?\"<>|]+", "_", folder_name)
    folder_name = " ".join(folder_name.split())
    folder_name = folder_name.strip(" .")

    if not folder_name:
        raise ValueError("Название папки не может быть пустым")
    if folder_name in {".", ".."}:
        raise ValueError("Такое название папки использовать нельзя")
    if folder_name == "__MACOSX":
        raise ValueError("Такое название папки зарезервировано")

    return folder_name[:60]


def create_folder_in_cards(parent_path, folder_name):
    """Создаёт новую папку внутри cards и не даёт выйти за пределы cards."""
    parent_path = ensure_safe_cards_path(parent_path)
    safe_name = sanitize_folder_name(folder_name)
    target_path = ensure_safe_cards_path(os.path.join(parent_path, safe_name))

    if os.path.exists(target_path):
        raise ValueError("Такая папка уже существует")

    os.makedirs(target_path, exist_ok=False)
    return target_path



def get_folder_delete_summary(folder_path):
    """Считает, что будет удалено вместе с папкой."""
    folder_path = ensure_safe_cards_path(folder_path)
    files_count = 0
    dirs_count = 0

    for _root, dirs, files in os.walk(folder_path):
        dirs[:] = [d for d in dirs if d != "__MACOSX"]
        dirs_count += len(dirs)
        files_count += len(files)

    return files_count, dirs_count


def delete_folder_from_cards(folder_path):
    """Удаляет выбранную папку внутри cards вместе с содержимым."""
    folder_path = ensure_safe_cards_path(folder_path)

    if folder_path == BASE_CARDS_DIR:
        raise ValueError("Нельзя удалить корневую папку cards")
    if not os.path.isdir(folder_path):
        raise ValueError("Папка уже не найдена")

    shutil.rmtree(folder_path)
    return folder_path


def get_tech_card_format_help():
    """Возвращает инструкцию для администратора по оформлению техкарты."""
    return (
        "📘 Как правильно оформить техкарту\n\n"
        "Принимаются файлы только в формате .xlsx.\n"
        "Файл нужно отправлять именно как документ VK, не фотографией и не текстом.\n"
        "Бот читает таблицу по строке с заголовками: «Наименование», «Ед изм», «Нетто».\n\n"
        "Обязательная структура файла:\n"
        "1. Вверху укажите название позиции: в одной ячейке «Наименование», "
        "а в соседней справа — название техкарты.\n"
        "2. Ниже сделайте строку заголовков: «Наименование | Ед изм | Нетто».\n"
        "3. Под заголовками перечислите ингредиенты.\n"
        "4. В конце можно добавить строку «Выход» — бот покажет итоговый выход.\n\n"
        "Пример техкарты: Смузи клубника-банан 0,3\n\n"
        "Строка 1:\n"
        "Наименование | Смузи клубника-банан 0,3\n\n"
        "Строка 3 — заголовки таблицы:\n"
        "Наименование | Ед изм | Нетто\n\n"
        "Ниже ингредиенты:\n"
        "Банан | г | 80\n"
        "Клубника | г | 70\n"
        "Сок яблочный | мл | 150\n"
        "Сироп сахарный | мл | 10\n"
        "Выход | мл | 300\n\n"
        "Важно:\n"
        "• не объединяйте ячейки в строке заголовков таблицы;\n"
        "• колонка «Наименование» должна быть именно так подписана;\n"
        "• колонка «Ед изм» нужна для единиц измерения;\n"
        "• колонка «Нетто» нужна для количества;\n"
        "• название файла может быть любым, но лучше назвать его как позицию;\n"
        "• принимается именно .xlsx;\n"
        "• ниже я прикреплю готовый .xlsx-файл для примера;\n"
        "• после загрузки нажмите «🔄 Обновить базу»."
    )




def ensure_example_tech_card_file():
    """Возвращает пример техкарты .xlsx из отдельной папки tech_card_examples.

    Папка создаётся автоматически. Если старый пример лежит рядом с main.py,
    бот перенесёт его в tech_card_examples, чтобы не хранить пример в корне проекта.
    """
    os.makedirs(EXAMPLE_FILES_DIR, exist_ok=True)

    example_path = os.path.join(EXAMPLE_FILES_DIR, EXAMPLE_TECH_CARD_FILENAME)
    if os.path.exists(example_path):
        return example_path

    # Мягкая миграция со старой схемы: пример раньше лежал рядом с main.py.
    old_preferred_path = os.path.join(BASE_DIR, EXAMPLE_TECH_CARD_FILENAME)
    if os.path.exists(old_preferred_path):
        shutil.copy2(old_preferred_path, example_path)
        return example_path

    old_generic_path = os.path.join(BASE_DIR, "primer_tehkarty.xlsx")
    if os.path.exists(old_generic_path):
        shutil.copy2(old_generic_path, example_path)
        return example_path

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as error:
        raise RuntimeError(
            "Не получилось создать пример .xlsx. Установите openpyxl: pip install openpyxl"
        ) from error

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Пример техкарты"

    rows = [
        ["Наименование", "Смузи клубника-банан 0,3", ""],
        ["", "", ""],
        ["Наименование", "Ед изм", "Нетто"],
        ["Банан", "г", 80],
        ["Клубника", "г", 70],
        ["Сок яблочный", "мл", 150],
        ["Сироп сахарный", "мл", 10],
        ["Выход", "мл", 300],
    ]

    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="FFD966")
    bold_font = Font(bold=True)

    for cell in sheet[1]:
        cell.font = bold_font
        cell.fill = header_fill

    for cell in sheet[3]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_len + 3, 12), 35)

    workbook.save(example_path)
    return example_path

def send_tech_card_format_help(vk, upload, user_id):
    """Отправляет инструкцию по оформлению техкарты и прикрепляет пример .xlsx."""
    send_msg(vk, user_id, get_tech_card_format_help())

    try:
        example_path = ensure_example_tech_card_file()
        document = upload.document_message(
            doc=example_path,
            title="Пример техкарты.xlsx",
            peer_id=user_id,
        )
        doc_info = document.get("doc", {})
        attachment = f"doc{doc_info['owner_id']}_{doc_info['id']}"
        send_msg(
            vk,
            user_id,
            "📎 Прикрепил пример техкарты в формате .xlsx. Можно скачать файл, заменить название и ингредиенты, а потом загрузить через админку.",
            keyboard=create_admin_keyboard(),
            attachment=attachment,
        )
    except Exception as error:
        print(f"Ошибка отправки примера техкарты: {error}")
        send_msg(
            vk,
            user_id,
            f"Инструкцию отправил, но пример .xlsx из папки tech_card_examples прикрепить не получилось: {error}",
            keyboard=create_admin_keyboard(),
        )


def send_example_tech_card_file(vk, upload, user_id, keyboard=None):
    """Прикрепляет администратору пример .xlsx без длинного пояснения."""
    try:
        example_path = ensure_example_tech_card_file()
        document = upload.document_message(
            doc=example_path,
            title="Пример техкарты.xlsx",
            peer_id=user_id,
        )
        doc_info = document.get("doc", {})
        attachment = f"doc{doc_info['owner_id']}_{doc_info['id']}"
        send_msg(
            vk,
            user_id,
            "📎 Пример техкарты в формате .xlsx",
            keyboard=keyboard,
            attachment=attachment,
        )
    except Exception as error:
        print(f"Ошибка отправки примера техкарты: {error}")
        send_msg(
            vk,
            user_id,
            f"Не получилось прикрепить пример .xlsx из папки tech_card_examples: {error}",
            keyboard=keyboard,
        )

def is_safe_cards_path(path):
    """Проверяет, что путь находится строго внутри папки cards."""
    abs_path = os.path.abspath(path)
    return abs_path == BASE_CARDS_DIR or abs_path.startswith(BASE_CARDS_DIR + os.sep)


def ensure_safe_cards_path(path):
    """Возвращает абсолютный безопасный путь внутри cards или выбрасывает ошибку."""
    abs_path = os.path.abspath(path)
    if not is_safe_cards_path(abs_path):
        raise ValueError("Запрещён путь вне папки cards")
    return abs_path


def get_card_files_in_folder(folder_path):
    """Возвращает Excel/CSV-файлы в выбранной папке."""
    folder_path = ensure_safe_cards_path(folder_path)
    if not os.path.exists(folder_path):
        return []
    files = []
    for filename in sorted(os.listdir(folder_path)):
        file_path = os.path.join(folder_path, filename)
        ext = os.path.splitext(filename)[1].lower()
        if os.path.isfile(file_path) and ext in ALLOWED_CARD_EXTENSIONS:
            files.append(filename)
    return files




def parse_file_selection(text, total_count):
    """Разбирает номера файлов для массового удаления.

    Поддерживает форматы:
    - 2
    - 1, 3, 5
    - 1 3 5
    - 1-4
    - все
    """
    text = (text or "").strip().lower().replace("ё", "е")
    if text in {"все", "all", "*"}:
        return list(range(total_count))

    indexes = set()
    parts = re.split(r"[\s,;]+", text)

    for part in parts:
        part = part.strip()
        if not part:
            continue

        if "-" in part:
            start_raw, end_raw = part.split("-", 1)
            if not start_raw.isdigit() or not end_raw.isdigit():
                raise ValueError("Используйте номера файлов, например: 1, 3, 5 или 1-4")

            start = int(start_raw)
            end = int(end_raw)
            if start > end:
                start, end = end, start

            for number in range(start, end + 1):
                if number < 1 or number > total_count:
                    raise ValueError(f"Номера должны быть от 1 до {total_count}")
                indexes.add(number - 1)
            continue

        if not part.isdigit():
            raise ValueError("Используйте номера файлов, например: 1, 3, 5 или 1-4")

        number = int(part)
        if number < 1 or number > total_count:
            raise ValueError(f"Номера должны быть от 1 до {total_count}")
        indexes.add(number - 1)

    if not indexes:
        raise ValueError("Напишите хотя бы один номер файла.")

    return sorted(indexes)


def get_folder_tree_lines(root="cards"):
    """Формирует текстовое дерево папок cards."""
    root = ensure_safe_cards_path(root)
    lines = []
    if not os.path.exists(root):
        return ["Папка cards пока не найдена."]

    for current_root, dirs, _files in os.walk(root):
        dirs[:] = sorted(d for d in dirs if d != "__MACOSX")
        level = 0 if current_root == root else len(os.path.relpath(current_root, root).split(os.sep))
        indent = "  " * level
        folder_name = os.path.basename(current_root)
        if current_root == root:
            lines.append("cards/")
        else:
            lines.append(f"{indent}└─ {folder_name}/")
    return lines


def create_cards_zip_archive():
    """Создаёт ZIP-архив всех техкарт из cards с сохранением структуры папок."""
    cards_dir = ensure_safe_cards_path("cards")
    if not os.path.isdir(cards_dir):
        raise ValueError("Папка cards не найдена")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    temp_dir = tempfile.mkdtemp(prefix="vk_tech_cards_")
    zip_path = os.path.join(temp_dir, f"tech_cards_{timestamp}.zip")

    added_files = 0
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for current_root, dirs, files in os.walk(cards_dir):
            dirs[:] = sorted(d for d in dirs if d != "__MACOSX")
            for filename in sorted(files):
                if filename.startswith("._") or filename == ".DS_Store":
                    continue

                file_path = os.path.join(current_root, filename)
                if not os.path.isfile(file_path):
                    continue

                # В архив кладём путь начиная с cards/, чтобы сохранились все директории.
                arcname = os.path.relpath(file_path, BASE_DIR)
                archive.write(file_path, arcname)
                added_files += 1

    if added_files == 0:
        try:
            os.remove(zip_path)
            os.rmdir(temp_dir)
        except OSError:
            pass
        raise ValueError("В папке cards нет файлов для архивации")

    return zip_path, added_files


def send_cards_zip_archive(vk, upload, user_id):
    """Отправляет администратору ZIP-архив всех техкарт."""
    zip_path = None
    temp_dir = None
    try:
        zip_path, files_count = create_cards_zip_archive()
        temp_dir = os.path.dirname(zip_path)

        document = upload.document_message(
            doc=zip_path,
            title="Все техкарты.zip",
            peer_id=user_id,
        )
        doc_info = document.get("doc", {})
        attachment = f"doc{doc_info['owner_id']}_{doc_info['id']}"
        send_msg(
            vk,
            user_id,
            f"📦 Архив техкарт готов. Файлов внутри: {files_count}.",
            keyboard=create_admin_keyboard(),
            attachment=attachment,
        )
    except Exception as error:
        print(f"Ошибка отправки ZIP-архива техкарт: {error}")
        send_msg(
            vk,
            user_id,
            f"Не получилось отправить ZIP-архив: {error}",
            keyboard=create_admin_keyboard(),
        )
    finally:
        if zip_path and os.path.exists(zip_path):
            try:
                os.remove(zip_path)
            except OSError:
                pass
        if temp_dir and os.path.isdir(temp_dir):
            try:
                os.rmdir(temp_dir)
            except OSError:
                pass


def _normalize_vk_doc_ref(doc_ref):
    """Приводит ссылку на документ VK к формату owner_id_doc_id[_access_key]."""
    if not doc_ref:
        return None

    doc_ref = str(doc_ref).strip()
    if doc_ref.startswith("doc"):
        doc_ref = doc_ref[3:]

    return doc_ref or None


def _get_doc_by_ref(vk, doc_ref):
    """Безопасно получает документ VK по строковой ссылке."""
    doc_ref = _normalize_vk_doc_ref(doc_ref)
    if not doc_ref:
        return None

    try:
        docs = vk.docs.getById(docs=doc_ref)
    except Exception as error:
        print(f"Не удалось получить документ VK {doc_ref}: {error}")
        return None

    if docs:
        return docs[0]
    return None


def get_event_docs_info(vk, event):
    """Достаёт все документы из сообщения VK LongPoll.

    Админ может прислать сразу несколько .xlsx-файлов одним сообщением.
    VK может отдавать вложения в разных форматах. Для документов из личных
    сообщений часто нужен access_key, иначе docs.getById не возвращает URL
    или падает с ошибкой доступа. Поэтому собираем все возможные варианты.
    """
    attachments = getattr(event, "attachments", {}) or {}
    checked_refs = set()
    docs_found = []
    seen_docs = set()

    def add_doc(doc):
        if not doc:
            return

        owner_id = doc.get("owner_id")
        doc_id = doc.get("id")
        url = doc.get("url")
        title = doc.get("title")
        unique_key = (owner_id, doc_id, url, title)
        if unique_key in seen_docs:
            return

        seen_docs.add(unique_key)
        docs_found.append(doc)

    def try_doc_ref(doc_ref, access_key=None):
        doc_ref = _normalize_vk_doc_ref(doc_ref)
        if not doc_ref:
            return None

        refs = []
        if access_key and doc_ref.count("_") == 1:
            refs.append(f"{doc_ref}_{access_key}")
        refs.append(doc_ref)

        for ref in refs:
            if ref in checked_refs:
                continue
            checked_refs.add(ref)
            doc = _get_doc_by_ref(vk, ref)
            if doc and doc.get("url"):
                return doc
        return None

    # Формат LongPoll: attach1_type=doc, attach1=owner_id_doc_id,
    # attach2_type=doc, attach2=owner_id_doc_id и т.д.
    # Иногда отдельно приходит attach1_access_key.
    doc_prefixes = []
    for key, value in attachments.items():
        if key.endswith("_type") and value == "doc":
            doc_prefixes.append(key[: -len("_type")])

    doc_prefixes.sort(
        key=lambda prefix: int(prefix.replace("attach", ""))
        if prefix.replace("attach", "").isdigit()
        else 9999
    )

    for prefix in doc_prefixes:
        doc = try_doc_ref(
            attachments.get(prefix),
            attachments.get(f"{prefix}_access_key"),
        )
        add_doc(doc)

    # Запасной вариант: в некоторых событиях значение уже приходит как doc123_456.
    for value in attachments.values():
        if isinstance(value, str) and "_" in value:
            doc = try_doc_ref(value)
            add_doc(doc)

    # Ещё один запасной вариант: перечитываем сообщение через messages.getById.
    # Там вложения часто приходят структурированно: {'type': 'doc', 'doc': {...}}.
    message_id = getattr(event, "message_id", None)
    if message_id:
        try:
            response = vk.messages.getById(message_ids=message_id)
            items = response.get("items", []) if isinstance(response, dict) else []
        except Exception as error:
            print(f"Не удалось перечитать сообщение {message_id}: {error}")
            items = []

        for item in items:
            for attachment in item.get("attachments", []):
                if attachment.get("type") != "doc":
                    continue

                doc = attachment.get("doc") or {}
                if doc.get("url"):
                    add_doc(doc)
                    continue

                owner_id = doc.get("owner_id")
                doc_id = doc.get("id")
                access_key = doc.get("access_key")
                if owner_id and doc_id:
                    found_doc = try_doc_ref(f"{owner_id}_{doc_id}", access_key)
                    add_doc(found_doc)

    return docs_found


def get_event_doc_info(vk, event):
    """Оставлено для совместимости: возвращает первый документ из сообщения."""
    docs = get_event_docs_info(vk, event)
    return docs[0] if docs else None

def download_vk_doc_to_folder(doc_info, folder_path):
    """Скачивает документ VK в выбранную папку cards."""
    folder_path = ensure_safe_cards_path(folder_path)
    title = sanitize_filename(doc_info.get("title", "tech_card.xlsx"))
    ext = os.path.splitext(title)[1].lower()

    if ext not in ALLOWED_CARD_EXTENSIONS:
        raise ValueError("Можно загружать только файлы .xlsx")

    os.makedirs(folder_path, exist_ok=True)
    target_path = ensure_safe_cards_path(os.path.join(folder_path, title))

    if os.path.exists(target_path):
        name, ext = os.path.splitext(title)
        counter = 2
        while os.path.exists(target_path):
            target_path = ensure_safe_cards_path(os.path.join(folder_path, f"{name}_{counter}{ext}"))
            counter += 1

    url = doc_info.get("url")
    if not url:
        raise ValueError("У документа нет ссылки для скачивания")

    response = requests.get(
        url,
        timeout=30,
        stream=True,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    response.raise_for_status()

    expected_size = int(response.headers.get("content-length") or 0)
    if expected_size and expected_size > MAX_UPLOAD_SIZE:
        raise ValueError(f"Файл слишком большой. Максимум: {MAX_UPLOAD_SIZE // (1024 * 1024)} МБ")

    downloaded = 0
    with open(target_path, "wb") as file:
        for chunk in response.iter_content(chunk_size=1024 * 1024):
            if not chunk:
                continue
            downloaded += len(chunk)
            if downloaded > MAX_UPLOAD_SIZE:
                file.close()
                try:
                    os.remove(target_path)
                except OSError:
                    pass
                raise ValueError(f"Файл слишком большой. Максимум: {MAX_UPLOAD_SIZE // (1024 * 1024)} МБ")
            file.write(chunk)

    return target_path


def show_admin_menu(vk, user_id, admin_ids):
    """Открывает меню администратора."""
    if not is_admin(user_id, admin_ids):
        send_msg(
            vk,
            user_id,
            "У вас пока нет прав администратора.",
            keyboard=create_start_keyboard_for_user(user_id, admin_ids),
        )
        return

    admin_states.pop(user_id, None)
    send_msg(
        vk,
        user_id,
        "⚙️ Административная панель\n\n"
        "Здесь можно добавлять и удалять техкарты, создавать и удалять папки, "
        "скачивать все техкарты ZIP-архивом, а также смотреть пример оформления файла .xlsx.",
        keyboard=create_admin_keyboard(),
    )


def start_admin_folder_choice(vk, user_id, mode):
    """Запускает выбор папки для добавления/удаления файла, создания или удаления папки."""
    admin_states[user_id] = {"mode": mode, "path": "cards"}
    folders = get_admin_subfolders("cards")
    actions = {
        "upload": "загрузки новой техкарты",
        "delete": "удаления техкарты",
        "create": "создания новой папки",
        "delete_dir": "удаления папки",
    }
    action = actions.get(mode, "действия")

    extra_text = ""
    if mode == "create":
        extra_text = "\n\nМожно создать папку прямо в cards или зайти внутрь нужного раздела."
    elif mode == "delete_dir":
        extra_text = (
            "\n\nЗайдите в ненужную папку и нажмите «🗑 Удалить эту папку». "
            "Папка удалится вместе со всеми файлами внутри после подтверждения."
        )
    elif mode == "upload":
        extra_text = (
            "\n\nПеред загрузкой .xlsx-файла проверьте формат через кнопку «📘 Как оформить техкарту» "
            "в админке."
        )

    send_msg(
        vk,
        user_id,
        f"Выберите папку для {action}:{extra_text}",
        keyboard=create_admin_folders_keyboard(folders, "cards", mode),
    )


def handle_admin_state(vk, upload, user_id, text, event, admin_ids):
    """Обрабатывает многошаговые действия администратора."""
    state = admin_states.get(user_id)
    if not state:
        return False

    mode = state.get("mode")
    current_path = state.get("path", "cards")

    if text == "⚙️ Админка":
        show_admin_menu(vk, user_id, admin_ids)
        return True

    if text == "⬅️ Админ назад":
        if current_path == "cards":
            show_admin_menu(vk, user_id, admin_ids)
        else:
            new_path = os.path.dirname(current_path)
            state["path"] = new_path
            folders = get_admin_subfolders(new_path)
            send_msg(
                vk,
                user_id,
                f"Текущая папка: {new_path}",
                keyboard=create_admin_folders_keyboard(folders, new_path, mode),
            )
        return True

    # Навигация по папкам внутри админки.
    # Сравниваем и полное имя, и первые 40 символов: VK-кнопка может обрезать длинное название.
    for folder in get_admin_subfolders(current_path):
        folder_label = folder[:40]
        if text.lower() in {folder.lower(), folder_label.lower()}:
            new_path = os.path.join(current_path, folder)
            state["path"] = new_path
            folders = get_admin_subfolders(new_path)
            send_msg(
                vk,
                user_id,
                f"Текущая папка: {new_path}",
                keyboard=create_admin_folders_keyboard(folders, new_path, mode),
            )
            return True

    if mode == "create":
        if text == "➕ Создать здесь":
            state["mode"] = "create_wait_name"
            send_msg(
                vk,
                user_id,
                f"Текущая папка: {current_path}\n\n"
                "Напишите название новой папки. Например: Новые напитки",
                keyboard=json.dumps(
                    {
                        "one_time": False,
                        "buttons": [
                            [
                                {
                                    "action": {"type": "text", "label": "⬅️ Админ назад"},
                                    "color": "secondary",
                                },
                                {
                                    "action": {"type": "text", "label": "⚙️ Админка"},
                                    "color": "secondary",
                                },
                            ]
                        ],
                    },
                    ensure_ascii=False,
                ),
            )
            return True

    if mode == "create_wait_name":
        try:
            created_path = create_folder_in_cards(current_path, text)
        except Exception as error:
            send_msg(
                vk,
                user_id,
                f"Не получилось создать папку: {error}\n\n"
                "Напишите другое название или вернитесь в админку.",
            )
            return True

        admin_states.pop(user_id, None)
        send_msg(
            vk,
            user_id,
            f"✅ Папка создана:\n{created_path}\n\n"
            "Теперь можно зайти в «➕ Добавить техкарту» и загрузить файл в эту папку.",
            keyboard=create_admin_keyboard(),
        )
        return True

    if mode == "delete_dir":
        if text == "🗑 Удалить эту папку":
            if current_path == "cards":
                send_msg(vk, user_id, "Нельзя удалить корневую папку cards.")
                return True

            try:
                files_count, dirs_count = get_folder_delete_summary(current_path)
                folder_name = os.path.basename(current_path)
            except Exception as error:
                send_msg(vk, user_id, f"Не получилось проверить папку: {error}")
                return True

            state["mode"] = "delete_dir_confirm"
            state["delete_dir_path"] = current_path
            send_msg(
                vk,
                user_id,
                f"Вы точно хотите удалить папку?\n\n"
                f"📁 {folder_name}\n"
                f"Внутри будет удалено: файлов — {files_count}, подпапок — {dirs_count}.\n\n"
                "Это действие нельзя отменить.",
                keyboard=json.dumps(
                    {
                        "one_time": False,
                        "buttons": [
                            [
                                {
                                    "action": {"type": "text", "label": "✅ Да, удалить папку"},
                                    "color": "negative",
                                }
                            ],
                            [
                                {
                                    "action": {"type": "text", "label": "⚙️ Админка"},
                                    "color": "secondary",
                                }
                            ],
                        ],
                    },
                    ensure_ascii=False,
                ),
            )
            return True

    if mode == "delete_dir_confirm":
        if text != "✅ Да, удалить папку":
            admin_states.pop(user_id, None)
            send_msg(vk, user_id, "Удаление папки отменено.", keyboard=create_admin_keyboard())
            return True

        target_path = state.get("delete_dir_path")
        try:
            deleted_path = delete_folder_from_cards(target_path)
            message = (
                f"✅ Папка удалена:\n{os.path.basename(deleted_path)}\n\n"
                "Нажмите «🔄 Обновить базу», чтобы бот перечитал техкарты."
            )
        except Exception as error:
            message = f"Не получилось удалить папку: {error}"

        admin_states.pop(user_id, None)
        send_msg(vk, user_id, message, keyboard=create_admin_keyboard())
        return True

    if mode == "upload":
        if text == "✅ Выбрать эту папку":
            state["mode"] = "upload_wait_file"
            upload_keyboard = json.dumps(
                {
                    "one_time": False,
                    "buttons": [
                        [
                            {
                                "action": {"type": "text", "label": "⚙️ Админка"},
                                "color": "secondary",
                            }
                        ]
                    ],
                },
                ensure_ascii=False,
            )
            send_msg(
                vk,
                user_id,
                f"Папка выбрана: {current_path}\n\n"
                "Теперь отправьте сюда один или несколько файлов техкарт в формате .xlsx.",
                keyboard=upload_keyboard,
            )
            send_example_tech_card_file(vk, upload, user_id, keyboard=upload_keyboard)
            return True

    if mode == "upload_wait_file":
        docs_info = get_event_docs_info(vk, event)
        if not docs_info:
            send_msg(
                vk,
                user_id,
                "Пришлите один или несколько файлов документами VK в формате .xlsx.",
            )
            return True

        saved_paths = []
        errors = []
        for doc_info in docs_info:
            title = sanitize_filename(doc_info.get("title", "tech_card.xlsx"))
            try:
                saved_path = download_vk_doc_to_folder(doc_info, current_path)
                saved_paths.append(saved_path)
            except Exception as error:
                errors.append(f"{title}: {error}")

        if not saved_paths:
            send_msg(
                vk,
                user_id,
                "Не получилось сохранить ни один файл:\n" + "\n".join(errors[:10]),
            )
            return True

        admin_states.pop(user_id, None)

        saved_names = [os.path.basename(path) for path in saved_paths]
        message = (
            f"✅ Добавлено техкарт: {len(saved_paths)}\n"
            + "\n".join(f"• {name}" for name in saved_names[:20])
        )
        if len(saved_names) > 20:
            message += f"\n…и ещё {len(saved_names) - 20}"

        if errors:
            message += "\n\n⚠️ Не удалось сохранить некоторые файлы:\n"
            message += "\n".join(f"• {error}" for error in errors[:10])
            if len(errors) > 10:
                message += f"\n…и ещё ошибок: {len(errors) - 10}"

        message += "\n\nНажмите «🔄 Обновить базу», чтобы бот перечитал техкарты."

        send_msg(
            vk,
            user_id,
            message,
            keyboard=create_admin_keyboard(),
        )
        return True

    if mode == "delete":
        if text == "📄 Показать файлы здесь":
            files = get_card_files_in_folder(current_path)
            if not files:
                send_msg(
                    vk,
                    user_id,
                    "В этой папке нет файлов техкарт.",
                    keyboard=create_admin_folders_keyboard(
                        get_admin_subfolders(current_path), current_path, mode
                    ),
                )
                return True

            state["mode"] = "delete_choose_file"
            state["files"] = files
            lines = [f"{idx}. {filename}" for idx, filename in enumerate(files, 1)]
            message = (
                f"Файлы в папке {current_path}:\n\n"
                + "\n".join(lines)
                + "\n\nНапишите номера файлов для удаления:\n"
                "• один файл: 2\n"
                "• несколько файлов: 1, 3, 5\n"
                "• диапазон: 1-4\n"
                "• все файлы: все"
            )
            send_msg(
                vk,
                user_id,
                message,
                keyboard=json.dumps(
                    {
                        "one_time": False,
                        "buttons": [
                            [
                                {
                                    "action": {"type": "text", "label": "⬅️ Админ назад"},
                                    "color": "secondary",
                                },
                                {
                                    "action": {"type": "text", "label": "⚙️ Админка"},
                                    "color": "secondary",
                                },
                            ]
                        ],
                    },
                    ensure_ascii=False,
                ),
            )
            return True

    if mode == "delete_choose_file":
        files = state.get("files", [])
        try:
            selected_indexes = parse_file_selection(text, len(files))
        except ValueError as error:
            send_msg(vk, user_id, str(error))
            return True

        selected_files = [files[index] for index in selected_indexes]
        delete_paths = [
            ensure_safe_cards_path(os.path.join(current_path, filename))
            for filename in selected_files
        ]

        state["mode"] = "delete_confirm"
        state["delete_paths"] = delete_paths

        preview_lines = [f"• {filename}" for filename in selected_files[:30]]
        message = (
            f"Вы точно хотите удалить файлов: {len(selected_files)}?\n\n"
            + "\n".join(preview_lines)
        )
        if len(selected_files) > 30:
            message += f"\n…и ещё {len(selected_files) - 30}"

        send_msg(
            vk,
            user_id,
            message,
            keyboard=json.dumps(
                {
                    "one_time": False,
                    "buttons": [
                        [
                            {
                                "action": {"type": "text", "label": "✅ Да, удалить"},
                                "color": "negative",
                            }
                        ],
                        [
                            {
                                "action": {"type": "text", "label": "⬅️ Админ назад"},
                                "color": "secondary",
                            },
                            {
                                "action": {"type": "text", "label": "⚙️ Админка"},
                                "color": "secondary",
                            },
                        ],
                    ],
                },
                ensure_ascii=False,
            ),
        )
        return True

    if mode == "delete_confirm":
        if text != "✅ Да, удалить":
            send_msg(vk, user_id, "Удаление отменено.", keyboard=create_admin_keyboard())
            admin_states.pop(user_id, None)
            return True

        delete_paths = state.get("delete_paths")
        if not delete_paths:
            old_path = state.get("delete_path")
            delete_paths = [old_path] if old_path else []

        deleted = []
        missing = []
        errors = []

        for target_path in delete_paths:
            try:
                target_path = ensure_safe_cards_path(target_path)
                filename = os.path.basename(target_path)
                if os.path.exists(target_path):
                    os.remove(target_path)
                    deleted.append(filename)
                else:
                    missing.append(filename)
            except Exception as error:
                errors.append(f"{os.path.basename(str(target_path))}: {error}")

        if deleted:
            message = (
                f"✅ Удалено техкарт: {len(deleted)}\n"
                + "\n".join(f"• {filename}" for filename in deleted[:30])
            )
            if len(deleted) > 30:
                message += f"\n…и ещё {len(deleted) - 30}"
            message += "\n\nНажмите «🔄 Обновить базу», чтобы бот перечитал техкарты."
        else:
            message = "Файлы не были удалены."

        if missing:
            message += "\n\n⚠️ Уже не найдены:\n"
            message += "\n".join(f"• {filename}" for filename in missing[:10])
            if len(missing) > 10:
                message += f"\n…и ещё {len(missing) - 10}"

        if errors:
            message += "\n\n⚠️ Ошибки при удалении:\n"
            message += "\n".join(f"• {error}" for error in errors[:10])
            if len(errors) > 10:
                message += f"\n…и ещё ошибок: {len(errors) - 10}"

        admin_states.pop(user_id, None)
        send_msg(vk, user_id, message, keyboard=create_admin_keyboard())
        return True

    return False


# =========================
# ОБЩАЯ ИНФОРМАЦИЯ И PDF
# =========================


def create_info_keyboard(user_id=None, admin_ids=None):
    """Клавиатура после раздела общей информации."""
    buttons = [
        [
            {"action": {"type": "text", "label": "📖 Меню"}, "color": "positive"},
            {"action": {"type": "text", "label": "📝 Тесты"}, "color": "primary"},
        ],
        [{"action": {"type": "text", "label": "🏠 В начало"}, "color": "secondary"}],
    ]
    if user_id is not None and admin_ids is not None and is_admin(user_id, admin_ids):
        buttons.insert(
            1,
            [{"action": {"type": "text", "label": "⚙️ Админка"}, "color": "positive"}],
        )
    return json.dumps({"one_time": False, "buttons": buttons}, ensure_ascii=False)


def show_general_info(vk, upload, user_id, admin_ids):
    """Показывает описание бота и отправляет PDF для самостоятельного изучения."""
    user_test_sessions.pop(user_id, None)
    clear_nav(vk, user_id)

    message = (
        "ℹ️ Общая информация\n\n"
        "Привет! Я бот-помощник для изучения техкарт.\n\n"
        "Что я умею:\n"
        "• показывать техкарты напитков, блюд и десертов;\n"
        "• выводить состав позиции по названию;\n"
        "• помогать быстро переключаться между разделами меню;\n"
        "• запускать тесты для проверки знаний.\n\n"
        "Как пользоваться:\n"
        "1. Нажмите «📖 Меню».\n"
        "2. Выберите нужный раздел и позицию.\n"
        "3. Для самостоятельного изучения можно открыть PDF со всеми техкартами.\n\n"
        "PDF с техкартами отправлю следующим сообщением 📎"
    )
    send_msg(vk, user_id, message, keyboard=create_info_keyboard(user_id, admin_ids))

    # PDF со всеми техкартами храним вместе с файлами-примерами.
    # Ожидаемый путь: tech_card_examples/all_tech_cards.pdf
    os.makedirs(EXAMPLE_FILES_DIR, exist_ok=True)
    pdf_path = os.path.join(EXAMPLE_FILES_DIR, "all_tech_cards.pdf")
    if not os.path.exists(pdf_path):
        send_msg(
            vk,
            user_id,
            "PDF-файл пока не найден. Проверьте, что файл all_tech_cards.pdf лежит в папке tech_card_examples.",
            keyboard=create_info_keyboard(user_id, admin_ids),
        )
        return

    try:
        document = upload.document_message(
            doc=pdf_path,
            title="Все техкарты для изучения.pdf",
            peer_id=user_id,
        )
        doc_info = document.get("doc", {})
        attachment = f"doc{doc_info['owner_id']}_{doc_info['id']}"
        send_msg(
            vk,
            user_id,
            "📎 Все техкарты для самостоятельного изучения.",
            keyboard=create_info_keyboard(user_id, admin_ids),
            attachment=attachment,
        )
    except Exception as error:
        print(f"Ошибка отправки PDF: {error}")
        send_msg(
            vk,
            user_id,
            "Не получилось отправить PDF автоматически. Файл all_tech_cards.pdf лежит в папке tech_card_examples — его можно отправить вручную.",
            keyboard=create_info_keyboard(user_id, admin_ids),
        )


# =========================
# КЛАВИАТУРЫ ТЕХКАРТ
# =========================


def create_search_keyboard():
    """Клавиатура, которая появляется под текстом техкарты."""
    return json.dumps(
        {
            "buttons": [
                [
                    {
                        "action": {"type": "text", "label": "⬅️ К списку"},
                        "color": "primary",
                    }
                ],
                [
                    {
                        "action": {"type": "text", "label": "🏠 В начало"},
                        "color": "secondary",
                    }
                ],
            ]
        },
        ensure_ascii=False,
    )


def create_inline_drinks_keyboard(items):
    """Создает inline-клавиатуру максимум на 10 кнопок."""
    buttons = []
    current_row = []

    for item in items[:10]:
        button = {
            "action": {"type": "text", "label": item.capitalize()[:40]},
            "color": "primary",
        }
        current_row.append(button)
        if len(current_row) == 2:
            buttons.append(current_row)
            current_row = []

    if current_row:
        buttons.append(current_row)

    return json.dumps({"inline": True, "buttons": buttons}, ensure_ascii=False)


def _normalize_page(page, total_pages):
    """Не даёт странице выйти за допустимые границы."""
    if total_pages <= 1:
        return 0
    return max(0, min(page, total_pages - 1))


def get_cards_for_path(all_cards, current_path):
    """Возвращает техкарты, которые лежат в выбранном разделе."""
    rel_path = os.path.relpath(current_path, "cards")
    if rel_path == ".":
        return []

    rel_path_text = rel_path.replace(os.sep, " / ").lower()
    marker = f"📂 раздел: {rel_path_text}"

    return [name for name, content in all_cards.items() if marker in content.lower()]


def get_sibling_folders(current_path):
    """Возвращает соседние папки для быстрого перехода между разделами."""
    if current_path == "cards":
        return get_subfolders("cards")

    parent_path = os.path.dirname(current_path)
    if not parent_path or parent_path == current_path:
        return []

    return get_subfolders(parent_path)


def find_folder_path_by_name(folder_name, current_path="cards"):
    """Ищет папку по названию из любого места меню."""
    folder_name_lower = folder_name.strip().lower()
    search_roots = [current_path]

    parent_path = os.path.dirname(current_path)
    if parent_path and parent_path not in search_roots:
        search_roots.append(parent_path)

    if "cards" not in search_roots:
        search_roots.append("cards")

    for root in search_roots:
        for folder in get_subfolders(root):
            if folder.lower() == folder_name_lower:
                return os.path.join(root, folder)

    for root, dirs, _files in os.walk("cards"):
        dirs[:] = [d for d in dirs if d != "__MACOSX"]
        for folder in dirs:
            if folder.lower() == folder_name_lower:
                return os.path.join(root, folder)

    return None


def show_folder(vk, user_id, folder_path, folder_title, all_cards):
    """Открывает папку: показывает подпапки или список техкарт."""
    subfolders = get_subfolders(folder_path)
    if subfolders:
        send_msg(
            vk,
            user_id,
            f"📁 Внутри {folder_title}:",
            keyboard=create_folders_keyboard(subfolders),
        )
        return

    items = get_cards_for_path(all_cards, folder_path)
    nav_folders = get_sibling_folders(folder_path)
    send_cards_list(vk, user_id, f"✨ Раздел: {folder_title}", items, nav_folders)


def send_cards_list(vk, user_id, title, items, nav_folders=None, page=0):
    """Показывает ассортимент с учетом лимита VK."""
    nav_folders = nav_folders or []

    if not items:
        send_msg(
            vk,
            user_id,
            "В этом разделе техкарты не найдены.",
            keyboard=create_folders_keyboard(nav_folders),
        )
        return

    sorted_items = sorted(items)

    if len(sorted_items) <= 10:
        message = (
            f"{title}\n"
            f"Всего позиций: {len(sorted_items)}.\n\n"
            "Выберите позицию кнопкой ниже 👇"
        )
        # Inline-кнопки с позициями остаются прямо под сообщением.
        # Отдельным сообщением возвращаем обычную навигационную клавиатуру,
        # чтобы всегда работали «⬅️ Назад», «🏠 В начало» и переходы в соседние папки.
        send_msg(vk, user_id, message, keyboard=create_inline_drinks_keyboard(sorted_items))
        send_msg(
            vk,
            user_id,
            "Навигация по разделам:",
            keyboard=create_folders_keyboard(nav_folders),
        )
        return

    lines = [f"{idx}. {name.capitalize()}" for idx, name in enumerate(sorted_items, 1)]
    header = f"{title}\nВсего позиций: {len(sorted_items)}.\n\nАссортимент:"
    blocks = _split_text_blocks(lines, header=header)

    for block in blocks[:-1]:
        send_msg(vk, user_id, block)

    last_block = f"{blocks[-1]}\n\nЧтобы открыть техкарту, напишите название позиции из списка."
    send_msg(vk, user_id, last_block, keyboard=create_folders_keyboard(nav_folders))


# =========================
# ТЕСТЫ
# =========================


def create_tests_menu_keyboard():
    """Клавиатура выбора раздела тестов."""
    return json.dumps(
        {
            "one_time": False,
            "buttons": [
                [{"action": {"type": "text", "label": "🥤 Тесты напитки"}, "color": "positive"}],
                [{"action": {"type": "text", "label": "🍰 Тесты десерты"}, "color": "positive"}],
                [{"action": {"type": "text", "label": "🏠 В начало"}, "color": "secondary"}],
            ],
        },
        ensure_ascii=False,
    )


def create_test_list_keyboard(tests):
    """Кнопки со списком тестов."""
    buttons = []
    for test in tests[:8]:
        buttons.append(
            [{"action": {"type": "text", "label": test["title"][:40]}, "color": "primary"}]
        )

    buttons.append(
        [
            {"action": {"type": "text", "label": "⬅️ К тестам"}, "color": "secondary"},
            {"action": {"type": "text", "label": "🏠 В начало"}, "color": "secondary"},
        ]
    )
    return json.dumps({"one_time": False, "buttons": buttons}, ensure_ascii=False)


def create_answer_keyboard(options):
    """Inline-кнопки вариантов ответа."""
    buttons = []
    for row_options in _chunked(options, 2):
        buttons.append(
            [
                {"action": {"type": "text", "label": option[:40]}, "color": "primary"}
                for option in row_options
            ]
        )
    return json.dumps({"inline": True, "buttons": buttons}, ensure_ascii=False)


def show_tests_menu(vk, user_id):
    """Открывает главный раздел тестов из любого места бота."""
    user_test_sessions.pop(user_id, None)
    clear_nav(vk, user_id)
    send_msg(vk, user_id, "📝 Выберите раздел тестов:", keyboard=create_tests_menu_keyboard())


def show_tests_list(vk, user_id, tests, title):
    """Показывает список доступных тестов."""
    user_test_sessions.pop(user_id, None)
    message = f"{title}\n\nВыберите тест 👇"
    send_msg(vk, user_id, message, keyboard=create_test_list_keyboard(tests))


def find_test_by_title(text):
    """Ищет тест по названию кнопки."""
    text_lower = text.strip().lower()
    for test in DRINK_TESTS + DESSERT_TESTS:
        if test["title"].lower() == text_lower:
            return test
    return None


def start_test(vk, user_id, test):
    """Запускает выбранный тест."""
    user_test_sessions[user_id] = {"test": test, "question_index": 0, "score": 0}
    send_test_question(vk, user_id)


def send_test_question(vk, user_id):
    """Отправляет текущий вопрос теста."""
    session = user_test_sessions.get(user_id)
    if not session:
        show_tests_menu(vk, user_id)
        return

    test = session["test"]
    question_index = session["question_index"]
    questions = test["questions"]

    if question_index >= len(questions):
        score = session["score"]
        total = len(questions)
        user_test_sessions.pop(user_id, None)
        send_msg(
            vk,
            user_id,
            f"✅ Тест завершён!\nВаш результат: {score} из {total}.",
            keyboard=create_tests_menu_keyboard(),
        )
        return

    question = questions[question_index]
    message = (
        f"{test['title']}\n"
        f"Вопрос {question_index + 1} из {len(questions)}\n\n"
        f"{question['question']}"
    )

    option_texts = question.get("option_texts")
    if option_texts:
        message += "\n\n" + "\n\n".join(option_texts)

    send_msg(vk, user_id, message, keyboard=create_answer_keyboard(question["options"]))


def handle_test_answer(vk, user_id, text):
    """Проверяет ответ пользователя и отправляет следующий вопрос."""
    session = user_test_sessions.get(user_id)
    if not session:
        return False

    test = session["test"]
    question_index = session["question_index"]
    question = test["questions"][question_index]

    if text not in question["options"]:
        send_msg(
            vk,
            user_id,
            "Выберите один из вариантов ответа кнопкой ниже 👇",
            keyboard=create_answer_keyboard(question["options"]),
        )
        return True

    if text == question["answer"]:
        session["score"] += 1
        result_text = "✅ Верно!"
    else:
        result_text = f"❌ Неверно. Правильный ответ: {question['answer']}"

    session["question_index"] += 1
    send_msg(vk, user_id, result_text)
    send_test_question(vk, user_id)
    return True


# =========================
# ОСНОВНОЙ ЦИКЛ БОТА
# =========================


def main():
    # Запускаем бота из папки проекта, чтобы cards/ и PDF находились стабильно.
    os.chdir(BASE_DIR)

    vk_session = vk_api.VkApi(token=TOKEN)
    vk = vk_session.get_api()
    upload = VkUpload(vk_session)
    longpoll = VkLongPoll(vk_session)

    print("🤖 Бот запущен и сканирует файлы...")
    ALL_CARDS = load_all_cards()
    admin_ids = load_admin_ids()
    user_paths = {}

    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW and event.to_me:
            user_id = event.user_id
            text = event.text.strip()
            text_lower = text.lower()
            current_path = user_paths.get(user_id, "cards")

            # 1. ГЛОБАЛЬНЫЕ КОМАНДЫ
            if text_lower in ["привет", "начать", "🏠 в начало", "главное меню"]:
                user_paths[user_id] = "cards"
                user_pages[user_id] = 0
                user_test_sessions.pop(user_id, None)
                admin_states.pop(user_id, None)
                clear_nav(vk, user_id)
                send_msg(
                    vk,
                    user_id,
                    "Выберите раздел 👇",
                    keyboard=create_start_keyboard_for_user(user_id, admin_ids),
                )
                continue

            if text_lower == ADMIN_ACCESS_WORD:
                admin_states[user_id] = {"mode": "await_admin_secret"}
                send_msg(
                    vk,
                    user_id,
                    "🔐 Вход в административный режим.\n\n"
                    "Введите пароль администратора одним сообщением.",
                    keyboard=create_start_keyboard_for_user(user_id, admin_ids),
                )
                continue

            if admin_states.get(user_id, {}).get("mode") == "await_admin_secret":
                if text_lower.startswith("/админ пароль "):
                    entered_secret = text[len("/админ пароль ") :].strip()
                else:
                    entered_secret = text.strip()

                if entered_secret == ADMIN_SECRET:
                    add_admin(user_id, admin_ids)
                    admin_states.pop(user_id, None)
                    send_msg(
                        vk,
                        user_id,
                        "✅ Готово. Теперь вы администратор бота. В меню появилась кнопка «⚙️ Админка».",
                        keyboard=create_start_keyboard_for_user(user_id, admin_ids),
                    )
                else:
                    send_msg(
                        vk,
                        user_id,
                        "❌ Неверный код администратора. Попробуйте ещё раз или нажмите «🏠 В начало».",
                        keyboard=create_start_keyboard_for_user(user_id, admin_ids),
                    )
                continue

            if text_lower.startswith("/админ пароль "):
                entered_secret = text[len("/админ пароль ") :].strip()
                if entered_secret == ADMIN_SECRET:
                    add_admin(user_id, admin_ids)
                    admin_states.pop(user_id, None)
                    send_msg(
                        vk,
                        user_id,
                        "✅ Готово. Теперь вы администратор бота. В меню появилась кнопка «⚙️ Админка».",
                        keyboard=create_start_keyboard_for_user(user_id, admin_ids),
                    )
                else:
                    send_msg(vk, user_id, "❌ Неверный код администратора.")
                continue

            if text == "⚙️ Админка":
                show_admin_menu(vk, user_id, admin_ids)
                continue

            if text == "📖 Меню":
                user_paths[user_id] = "cards"
                user_pages[user_id] = 0
                user_test_sessions.pop(user_id, None)
                admin_states.pop(user_id, None)
                folders = get_subfolders("cards")
                send_msg(vk, user_id, "Категории:", keyboard=create_folders_keyboard(folders))
                continue

            if text == "📝 Тесты":
                admin_states.pop(user_id, None)
                show_tests_menu(vk, user_id)
                continue

            if text == "ℹ️ Общая информация":
                admin_states.pop(user_id, None)
                show_general_info(vk, upload, user_id, admin_ids)
                continue

            # 1.1. АДМИН-КОМАНДЫ
            if text in [
                "➕ Добавить техкарту",
                "🗑 Удалить техкарту",
                "📂 Создать папку",
                "🗂 Удалить папку",
                "📘 Как оформить техкарту",
                "📁 Папки техкарт",
                "📦 Скачать все техкарты ZIP",
                "🔄 Обновить базу",
                "👥 Список админов",
            ]:
                if not is_admin(user_id, admin_ids):
                    send_msg(vk, user_id, "Эта команда доступна только администратору.")
                    continue

                if text == "➕ Добавить техкарту":
                    start_admin_folder_choice(vk, user_id, "upload")
                    continue

                if text == "🗑 Удалить техкарту":
                    start_admin_folder_choice(vk, user_id, "delete")
                    continue

                if text == "📂 Создать папку":
                    start_admin_folder_choice(vk, user_id, "create")
                    continue

                if text == "🗂 Удалить папку":
                    start_admin_folder_choice(vk, user_id, "delete_dir")
                    continue

                if text == "📘 Как оформить техкарту":
                    send_tech_card_format_help(vk, upload, user_id)
                    continue

                if text == "📁 Папки техкарт":
                    lines = get_folder_tree_lines("cards")
                    blocks = _split_text_blocks(lines, header="📁 Структура папок техкарт:")
                    for block in blocks[:-1]:
                        send_msg(vk, user_id, block)
                    send_msg(vk, user_id, blocks[-1], keyboard=create_admin_keyboard())
                    continue

                if text in {"📦 Скачать все техкарты ZIP", "📦 Скачать техкарты ZIP"}:
                    send_msg(vk, user_id, "Собираю ZIP-архив техкарт, подождите немного...")
                    send_cards_zip_archive(vk, upload, user_id)
                    continue

                if text == "🔄 Обновить базу":
                    ALL_CARDS = load_all_cards()
                    send_msg(
                        vk,
                        user_id,
                        f"✅ База техкарт обновлена. Загружено техкарт: {len(ALL_CARDS)}.",
                        keyboard=create_admin_keyboard(),
                    )
                    continue

                if text == "👥 Список админов":
                    admin_list = "\n".join(f"• VK ID: {admin_id}" for admin_id in sorted(admin_ids))
                    send_msg(
                        vk,
                        user_id,
                        f"👥 Администраторы:\n{admin_list or 'Пока список пуст.'}",
                        keyboard=create_admin_keyboard(),
                    )
                    continue

            if handle_admin_state(vk, upload, user_id, text, event, admin_ids):
                continue

            if text == "⬅️ К тестам":
                show_tests_menu(vk, user_id)
                continue

            if text == "🥤 Тесты напитки":
                show_tests_list(vk, user_id, DRINK_TESTS, "🥤 Тесты по напиткам")
                continue

            if text == "🍰 Тесты десерты":
                show_tests_list(vk, user_id, DESSERT_TESTS, "🍰 Тесты по десертам")
                continue

            selected_test = find_test_by_title(text)
            if selected_test:
                start_test(vk, user_id, selected_test)
                continue

            if handle_test_answer(vk, user_id, text):
                continue

            # 2. НАЗАД И СПИСКИ
            if text == "⬅️ Назад":
                if current_path == "cards":
                    clear_nav(vk, user_id)
                    send_msg(
                        vk,
                        user_id,
                        "Главное меню:",
                        keyboard=create_start_keyboard_for_user(user_id, admin_ids),
                    )
                else:
                    parent_path = os.path.dirname(current_path)
                    if not parent_path or parent_path == current_path:
                        parent_path = "cards"

                    user_paths[user_id] = parent_path
                    user_pages[user_id] = 0

                    if parent_path == "cards":
                        send_msg(
                            vk,
                            user_id,
                            "Категории:",
                            keyboard=create_folders_keyboard(get_subfolders("cards")),
                        )
                    else:
                        show_folder(
                            vk,
                            user_id,
                            parent_path,
                            os.path.basename(parent_path),
                            ALL_CARDS,
                        )
                continue

            if text == "⬅️ К списку":
                items = get_cards_for_path(ALL_CARDS, current_path)
                section_name = os.path.basename(current_path)
                nav_folders = get_sibling_folders(current_path)
                user_pages[user_id] = 0
                send_cards_list(vk, user_id, f"📋 Список {section_name}:", items, nav_folders)
                continue

            # 4. НАВИГАЦИЯ ПО ПАПКАМ
            folder_path = find_folder_path_by_name(text, current_path)
            if folder_path:
                user_paths[user_id] = folder_path
                user_pages[user_id] = 0
                show_folder(vk, user_id, folder_path, os.path.basename(folder_path), ALL_CARDS)
                continue

            # 5. ПОИСК ТЕХКАРТЫ
            matches = [name for name in ALL_CARDS.keys() if name and text_lower == name.lower()]
            if not matches:
                matches = [name for name in ALL_CARDS.keys() if name and text_lower in name.lower()]

            if len(matches) == 1:
                send_msg(
                    vk,
                    user_id,
                    ALL_CARDS[matches[0]],
                    keyboard=create_search_keyboard(),
                    is_final=True,
                )
            elif len(matches) > 1:
                send_msg(
                    vk,
                    user_id,
                    "🔍 Уточните, что именно вы ищете:",
                    keyboard=create_inline_drinks_keyboard(matches),
                )


if __name__ == "__main__":
    main()
